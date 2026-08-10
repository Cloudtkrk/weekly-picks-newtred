#!/usr/bin/env python3
"""
TikTok Shop 週次おすすめ商品データ v1 (スコアリングなし版)
条件フィルタ + タグ付け + 売上順で整理して Excel / Discordテキストを出力
"""
import argparse
import datetime as dt
import sys

import os

import pandas as pd
import requests

sys.path.insert(0, ".")
from weekly_picks import (load_videos, load_products, tag_and_filter,
                          fmt_money, pct_rank, load_own_list, CONFIG)


def apply_own_list(tbl: pd.DataFrame, own_by_id: dict,
                   own_by_name: dict | None = None) -> tuple[pd.DataFrame, int]:
    """自社案件リストと突合し、一致した商品は遷移先をアフィリエイトリンクに
    差し替え + 🎁自社サンプル可バッジを付与。商品IDで一致が最優先、
    同名・別IDの再出品にも当たるよう正規化商品名でもフォールバック突合する"""
    import re
    from weekly_picks import norm_name
    if not own_by_id:
        return tbl, 0
    links, badges, n = [], [], 0
    for link, badge, name in zip(tbl["商品リンク"], tbl["バッジ"], tbl["商品名"]):
        m = re.search(r"/product/(\d+)", str(link))
        row = own_by_id.get(m.group(1)) if m else None
        if row is None and own_by_name:
            row = own_by_name.get(norm_name(str(name)))
        if row:
            n += 1
            links.append(row["アフィリエイトリンク"])
            if "🎁" not in badge:
                badge = (badge + "・" if badge else "") + "🎁自社サンプル可"
        else:
            links.append(link)
        badges.append(badge)
    tbl = tbl.copy()
    tbl["商品リンク"] = links
    tbl["バッジ"] = badges
    return tbl, n


def build_table(prod: pd.DataFrame, videos: pd.DataFrame,
                small_keys: set | None = None) -> pd.DataFrame:
    """商品×直近7日動画実績を突合して表示用テーブルを作る。
    small_keys: フォロワー少クリエイター動画エクスポートに売上が載っていた商品キー
    (=実績の少ないクリエイターでも売れている実証。🔰判定に使う)"""
    vc = ["recent_video_gmv_7d", "n_videos_7d", "n_new_posts",
          "median_gpm", "organic_gmv_share"]
    df = prod.merge(videos[["product_key"] + vc], on="product_key", how="left")
    df[vc] = df[vc].fillna(0)
    small_hit = df["product_key"].isin(small_keys or set())

    # 新指標: 参画ペース (人/月) = 参画数÷掲載月数 (最低1ヶ月)、1人あたりGMV = 30日売上÷参画数
    # 累積参画数は掲載期間で歪む (1年で150人と1ヶ月で150人は別物)。
    # 高GMVでも参画1000人超なら飽和していて新規の取り分が小さい
    months = ((pd.Timestamp.now() - df["listed_at"]).dt.days / 30).clip(lower=1)
    pace = df["n_creators"] / months                                # 掲載日欠損はNaN→"-"
    gmv_pc = (df["gmv_30d"] / df["n_creators"]).where(df["n_creators"] > 0)

    # バッジ (サンプル承認難易度)。NaNとの比較はFalseになり自動的に非該当
    # 🔰は「参画ペース×成立率×⭐」または「フォロワー少クリエイター動画で直近売上あり」
    c = CONFIG
    has_star = df["recent_video_gmv_7d"] > 0
    is_easy = ((pace >= c["easy_pace"]) & (df["creator_cvr_pct"] >= c["easy_cvr"]) & has_star) | small_hit
    is_gem = ~is_easy & (gmv_pc >= c["gem_gmv_per_creator"]) & (df["creator_cvr_pct"] >= c["gem_cvr"])
    is_own = df["product_name"].fillna("").apply(
        lambda s: any(b in s for b in c["own_sample_brands"]))

    def join_badges(e, g, o):
        b = []
        if e: b.append("🔰初心者でも狙いやすい")
        elif g: b.append("💎狙い目（実績者向け）")
        if o: b.append("🎁自社サンプル可")
        return "・".join(b)
    badge = [join_badges(*t) for t in zip(is_easy, is_gem, is_own)]

    def tags(r):
        t = []
        if r["is_campaign"]: t.append("🎫キャンペーン中")
        if r["is_seasonal"]: t.append("季節")
        if r["is_yakki"]: t.append("薬機法注意")
        if r["live_share"] >= 0.6: t.append("LIVE向き")
        elif r["live_share"] <= 0.2: t.append("動画向き")
        if r["recent_video_gmv_7d"] > 0 and r["organic_gmv_share"] >= 0.7:
            t.append("オーガニック")
        elif r["recent_video_gmv_7d"] > 0 and r["organic_gmv_share"] <= 0.3:
            t.append("広告主導")
        return "・".join(t)

    major = df["category"].astype(str).str.split(">").str[0].str.strip()
    out = pd.DataFrame({
        "おすすめ": has_star.map({True: "⭐", False: ""}),
        "バッジ": badge,
        "商品名": df["product_name"].str.slice(0, 50),
        "ジャンル": major.map(lambda x: CONFIG["genre_map"].get(x, "その他")),
        "大分類": major,
        "カテゴリ": df["category"].astype(str).str.split(">").str[-1].str.strip(),
        "30日売上": df["gmv_30d"].map(fmt_money),
        "成長率": df["growth_pct"].map(lambda v: f"{v:+.0f}%"),
        "直近7日動画売上": df["recent_video_gmv_7d"].map(fmt_money),
        "直近7日動画本数": df["n_videos_7d"].astype(int),
        "販売件数": df["units"].astype(int).map(lambda v: f"{v/10000:.1f}万件" if v >= 10000 else f"{v:,}件"),
        "単価": df["avg_price"].map(fmt_money),
        "報酬目安/件": (df["avg_price"] * df["commission_pct"] / 100).map(
            lambda v: f"約{v:,.0f}円" if v > 0 else "-"),
        "報酬率": df["commission_pct"].map(lambda v: f"{v:.0f}%" if v else "-"),
        "参画クリエイター数": df["n_creators"].astype(int),
        "参画ペース": pace.map(lambda v: f"{v:.0f}人/月" if v == v else "-"),
        "1人あたりGMV": gmv_pc.map(lambda v: fmt_money(v) if v == v else "-"),
        "成立率": df["creator_cvr_pct"].map(lambda v: f"{v:.0f}%"),
        "評価": df["rating"],
        "タグ": df.apply(tags, axis=1),
        "掲載日": df["listed_at"].dt.strftime("%Y-%m-%d"),
        "商品リンク": df["tiktok_link"],
        "画像": df["image_link"],
        "画像alt": df["image_link_alt"],
        "_gmv": df["gmv_30d"],
        "_growth": df["growth_pct"],
        "_listed": df["listed_at"],
        "_cvr": df["creator_cvr_pct"],
        "_gmv_pc": gmv_pc,
    })
    return out


# ============================================================
# Excel画像埋め込み
# ============================================================
IMG_CACHE = ".img_cache"
THUMB_PX = 72          # サムネイル一辺 (px)
ROW_PT = 58            # 行高 (pt) ≒ 76px

def _fetch_thumb(url: str) -> str | None:
    """画像をDL→サムネイル化してキャッシュ。失敗時None"""
    import hashlib
    from PIL import Image as PILImage
    os.makedirs(IMG_CACHE, exist_ok=True)
    key = hashlib.md5(url.encode()).hexdigest()
    path = os.path.join(IMG_CACHE, f"{key}.png")
    if os.path.exists(path):
        return path
    headers = {
        "User-Agent": ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"),
        "Accept": "image/webp,image/png,image/*,*/*;q=0.8",
        "Referer": "https://www.kalodata.com/",
    }
    try:
        r = requests.get(url, timeout=10, headers=headers)
        r.raise_for_status()
        if "image" not in r.headers.get("Content-Type", ""):
            raise ValueError(f"not an image: {r.headers.get('Content-Type')}")
        from io import BytesIO
        im = PILImage.open(BytesIO(r.content)).convert("RGB")
        im.thumbnail((THUMB_PX, THUMB_PX))
        im.save(path, "PNG")
        return path
    except Exception as e:
        print(f"[warn] 画像取得失敗: {url[:60]}... ({e})")
        return None


def embed_images(xlsx_path: str, sheet_names: list[str]):
    """各シートの「画像」列(URL)を実画像に置き換えて埋め込む"""
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    wb = load_workbook(xlsx_path)
    for sn in sheet_names:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        headers = {c.value: i + 1 for i, c in enumerate(ws[1])}
        img_col = headers.get("画像")
        if not img_col:
            continue
        col_letter = get_column_letter(img_col)
        ws.column_dimensions[col_letter].width = 11
        n_ok = 0
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=img_col)
            url = cell.value
            if not (isinstance(url, str) and url.startswith("http")):
                continue
            path = _fetch_thumb(url)
            if not path:
                alt_col = headers.get("画像alt")
                alt = ws.cell(row=row, column=alt_col).value if alt_col else None
                if isinstance(alt, str) and alt.startswith("http") and alt != url:
                    path = _fetch_thumb(alt)
            if path:
                cell.value = None      # 取得成功時のみURLを消して画像に置き換え
                img = XLImage(path)
                img.anchor = f"{col_letter}{row}"
                ws.add_image(img)
                ws.row_dimensions[row].height = ROW_PT
                n_ok += 1
            # 失敗時はURLをそのまま残す (クリックで画像を開ける)
        print(f"[info] {sn}: 画像 {n_ok}件 埋め込み")
    wb.save(xlsx_path)


def make_embed(r, kind: str) -> dict:
    """Webカード (report_html._card) と同じ情報構成にする:
    ⭐/バッジ → 商品名 → 30日売上・単価・報酬率の3項目 → タグ。
    販売件数・成立率・参画数・報酬目安・成長率・掲載日などの詳細指標は
    Webに出さない方針なのでDiscordにも出さない (Excel/data JSONにのみ残す)"""
    color = 0xE74C3C if kind == "hot" else 0x2ECC71
    desc = []
    badges = str(r.get("バッジ") or "")
    if badges:
        desc.append(badges.replace("・", "　"))       # Webのピル行に相当
    desc.append(f"30日売上 **{r['30日売上']}**｜単価 {r['単価']}｜報酬率※ **{r['報酬率']}**")
    if r["タグ"]:
        desc.append(f"🏷️ {r['タグ']}")
    star = "⭐ " if r.get("おすすめ") == "⭐" and kind == "hot" else ""
    embed = {
        "title": (star + str(r["商品名"]))[:100],
        "url": r["商品リンク"] if isinstance(r["商品リンク"], str) and r["商品リンク"].startswith("http") else None,
        "description": "\n".join(desc),
        "color": color,
    }
    if isinstance(r["画像"], str) and r["画像"].startswith("http"):
        embed["thumbnail"] = {"url": r["画像"]}
    return {k: v for k, v in embed.items() if v is not None}


def discord_header(today: str) -> str:
    """Discord冒頭メッセージ。バッジ凡例 + サイトへの導線 (URL・合言葉の入手先)"""
    site = CONFIG.get("site_url", "") or ""
    line = CONFIG.get("line_add_url", "") or ""
    nav = ""
    if site:
        nav = (f"\n**▼ 全件はこちら（毎週更新）**\n"
               f"🔥 売れ筋・🚀 新商品・🎁 おすすめ・📺 ライブ\n{site}\n")
        # Discordは登録済みクリエイターのみのチャンネルなので合言葉を直接記載する
        if CONFIG.get("site_passcode"):
            nav += (f"🔑 合言葉： **{CONFIG['site_passcode']}**\n"
                    f"-# 初回のみ入力が必要です（一度入れれば以降そのまま見られます）\n")
    return (
        f"# 📦 今週のおすすめ商品（{today}更新）\n"
        f"⭐ **今週売れてる**＝直近7日も動画で売れている（今から乗っても間に合う）\n"
        f"🔰 **初心者でも狙いやすい**＝承認ペースが速く成立率も高い、"
        f"またはフォロワーの少ないクリエイターでも売れている実績あり\n"
        f"🎁 **自社サンプル可**＝弊社経由でサンプルを渡せる商品（実績ゼロでもまずここから）\n"
        f"💎 **狙い目（実績者向け）**＝1人あたりの取り分が大きいが、承認には実績や投稿数が必要\n"
        f"{nav}"
        f"-# ※報酬率は取得時点の参考値です。実際の料率は各案件のアフィリエイトセンターで必ず確認してください\n"
        f"-# ⚠️ 薬機法注意タグの商品は投稿前に表現チェック相談へ")


def discord_footer() -> str:
    """末尾メッセージ (再度サイトへ誘導)"""
    site = CONFIG.get("site_url", "") or ""
    if not site:
        return ""
    return (f"-# ここに載せているのは各ジャンルの上位数件です。\n"
            f"**全件・ショップ別・タイムセール依頼はサイトから** 👉 {site}")


def post_discord_embeds(hot, new_tbl, webhook, top_per_cat, preview_path: str | None = None):
    """preview_path を渡すと送信せず、投稿予定の内容をテキストに書き出す"""
    import time
    preview_lines = []

    def send(content=None, embeds=None):
        if preview_path is not None:
            if content:
                preview_lines.append(content)
            for e in embeds or []:
                preview_lines.append(
                    f"　┌ {e.get('title','')}\n"
                    + "\n".join("　│ " + l for l in e.get("description", "").split("\n"))
                    + (f"\n　└ 🔗 {e['url']}" if e.get("url") else "\n　└"))
            preview_lines.append("")
            return
        payload = {}
        if content: payload["content"] = content
        if embeds: payload["embeds"] = embeds
        # Discord Webhookはレート制限が厳しい(概ね5リクエスト/2秒)ため、
        # 429はRetry-Afterに従って再試行し、送信間隔も空ける
        for _ in range(5):
            r = requests.post(webhook, json=payload, timeout=15)
            if r.status_code != 429:
                r.raise_for_status()
                time.sleep(0.6)
                return
            try:
                wait = float(r.headers.get("Retry-After") or r.json().get("retry_after", 2))
            except Exception:
                wait = 2.0
            time.sleep(wait + 0.5)
        r.raise_for_status()

    # 冒頭のバッジ凡例はWebの「自分に合う商品の選び方」ガイドとフッター注記に合わせる
    today = dt.date.today().strftime("%m/%d")
    send(content=discord_header(today))
    for label, tbl, kind in [("🔥 売れ筋", hot, "hot"), ("🚀 新商品", new_tbl, "challenge")]:
        for cat, g in tbl.groupby("ジャンル", sort=False):
            g = g.head(top_per_cat)
            embeds = [make_embed(r, kind) for _, r in g.iterrows()]
            # Discordは1メッセージ10 embedまで
            for i in range(0, len(embeds), 10):
                head = f"## {label}｜▼ {cat}" if i == 0 else None
                send(content=head, embeds=embeds[i:i+10])
    foot = discord_footer()
    if foot:
        send(content=foot)
    if preview_path is not None:
        with open(preview_path, "w", encoding="utf-8") as f:
            f.write("\n".join(preview_lines))


def autodetect_inputs(input_dir: str):
    """input_dir のKalodataエクスポートを自動判別して
    (売れ筋, 新商品, 動画, フォロワー少動画orNone) を返す。判別できなければエラーで停止。
    - Kalodata_Product_*.xlsx ×2: 「アップロード時間」が全行45日以内の方が新商品
    - Kalodata_Video_*.xlsx ×1〜2: 2つある場合は行数が多い方が通常動画、少ない方がフォロワー少動画"""
    import glob
    prods = sorted(glob.glob(os.path.join(input_dir, "Kalodata_Product_*.xlsx")))
    vids = sorted(glob.glob(os.path.join(input_dir, "Kalodata_Video_*.xlsx")))
    if len(prods) != 2:
        sys.exit(f"[error] --auto: {input_dir} に Kalodata_Product_*.xlsx が2ファイル必要です "
                 f"(検出 {len(prods)}件: {prods})")
    if not 1 <= len(vids) <= 2:
        sys.exit(f"[error] --auto: {input_dir} に Kalodata_Video_*.xlsx が1〜2ファイル必要です "
                 f"(検出 {len(vids)}件: {vids})")

    def all_recent(path: str, days: int = 45) -> bool:
        df = pd.read_excel(path)
        if "アップロード時間" not in df.columns:
            sys.exit(f"[error] --auto: {path} に「アップロード時間」列がありません")
        ts = pd.to_datetime(df["アップロード時間"], errors="coerce").dropna()
        if ts.empty:
            sys.exit(f"[error] --auto: {path} の「アップロード時間」を日付として読めません")
        return bool((ts >= pd.Timestamp.now() - pd.Timedelta(days=days)).all())

    recent = [all_recent(p) for p in prods]
    if recent == [True, False]:
        new_f, hot_f = prods
    elif recent == [False, True]:
        hot_f, new_f = prods
    else:
        sys.exit("[error] --auto: 商品2ファイルの売れ筋/新商品を判別できません "
                 f"(全行45日以内フラグ: {dict(zip(prods, recent))})。"
                 "--products/--new を明示指定してください")

    if len(vids) == 2:
        rows = [len(pd.read_excel(v)) for v in vids]
        if rows[0] == rows[1]:
            sys.exit(f"[error] --auto: 動画2ファイルが同じ行数({rows[0]})で判別できません。"
                     "--videos/--videos-small を明示指定してください")
        main_v, small_v = (vids[0], vids[1]) if rows[0] > rows[1] else (vids[1], vids[0])
    else:
        main_v, small_v = vids[0], None

    print(f"[info] --auto判別: 売れ筋={hot_f} / 新商品={new_f} / 動画={main_v}"
          + (f" / フォロワー少動画={small_v}" if small_v else ""))
    return hot_f, new_f, main_v, small_v


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--auto", metavar="DIR", default=None,
                    help="指定ディレクトリのKalodataエクスポートを自動判別 "
                         "(--products/--new/--videos/--videos-small の代わり)")
    ap.add_argument("--products", default=None)
    ap.add_argument("--new", default=None)
    ap.add_argument("--videos", default=None)
    ap.add_argument("--videos-small", default=None,
                    help="フォロワー少クリエイターの上位動画エクスポート (🔰判定の実証データ、任意)")
    ap.add_argument("--out", default=None)
    ap.add_argument("--post", action="store_true", help="Discordにembed形式で投稿")
    ap.add_argument("--preview-post", metavar="PATH", nargs="?", const="discord_preview.txt",
                    help="Discordに送らず、投稿予定の内容をテキストに書き出して確認する")
    ap.add_argument("--webhook", default=os.environ.get("DISCORD_WEBHOOK_URL", ""))
    ap.add_argument("--top-per-cat", type=int, default=5,
                    help="Discord投稿時の1カテゴリあたり上限件数")
    ap.add_argument("--images", action="store_true",
                    help="商品画像をダウンロードしてExcelに埋め込む")
    ap.add_argument("--html", action="store_true",
                    help="画像付きHTMLレポートも出力する")
    ap.add_argument("--html-embed", action="store_true",
                    help="画像をbase64でHTML内に埋め込む (どこで開いても表示される自己完結版)")
    ap.add_argument("--no-archive", action="store_true",
                    help="archive/<日付>/ スナップショットと data/<日付>.json を保存しない")
    args = ap.parse_args()

    if args.auto:
        args.products, args.new, args.videos, args.videos_small = autodetect_inputs(args.auto)
    elif not (args.products and args.new and args.videos):
        ap.error("--products/--new/--videos を指定するか --auto <DIR> を使ってください")

    videos = load_videos(args.videos)
    small_keys = set()
    if args.videos_small:
        vs = load_videos(args.videos_small)
        small_keys = set(vs.loc[vs["recent_video_gmv_7d"] > 0, "product_key"])
        print(f"[info] フォロワー少動画: 商品{len(small_keys)}件を🔰判定に使用")
    prod_main = tag_and_filter(load_products(args.products))
    prod_main = prod_main[prod_main["n_creators"] > CONFIG["hot_min_creators"]]  # 売れ筋のみ: 低参画=専売疑いを除外
    prod_new = tag_and_filter(load_products(args.new))
    # 新商品: 成長率マイナスと参画3人以下を除外
    prod_new = prod_new[(prod_new["growth_pct"] >= CONFIG["challenge_min_growth"])
                        & (prod_new["n_creators"] > CONFIG["challenge_min_creators"])]

    hot = build_table(prod_main, videos, small_keys)
    # 新商品: 掲載日数以内(データにあれば適用)
    new_tbl = build_table(prod_new, videos, small_keys)
    cutoff = pd.Timestamp.now() - pd.Timedelta(days=CONFIG["challenge_max_listed_days"])
    recent_mask = new_tbl["_listed"] >= cutoff
    if recent_mask.any():
        new_tbl = new_tbl[recent_mask]

    # 並び順: ジャンル固定順 → ⭐(直近7日動画売上あり)優先 → 乗りやすさ順。
    # 乗りやすさ = 成立率のパーセンタイル + 1人あたりGMVのパーセンタイル
    # (売上高順は廃止: 高GMVでも飽和商品は新規の取り分が小さい)
    genre_pos = {g: i for i, g in enumerate(CONFIG["genre_order"])}
    hot["_genre"] = hot["ジャンル"].map(genre_pos)
    hot["_star"] = (hot["おすすめ"] == "⭐").astype(int)
    hot["_ride"] = pct_rank(hot["_cvr"].fillna(0)) + pct_rank(hot["_gmv_pc"].fillna(0))
    hot = hot.sort_values(["_genre", "_star", "_ride"],
                          ascending=[True, False, False], kind="stable")
    hot = hot.groupby("ジャンル", sort=False).head(CONFIG["per_genre_limit"])
    # 新商品は件数制限なし・同じジャンル構成で成長率順のまま
    new_tbl["_genre"] = new_tbl["ジャンル"].map(genre_pos)
    new_tbl = new_tbl.sort_values(["_genre", "_growth"],
                                  ascending=[True, False], kind="stable")

    drop = ["_gmv", "_growth", "_listed", "_cvr", "_gmv_pc", "_genre", "_star", "_ride"]
    hot_gmv_raw = hot["_gmv"]
    new_gmv_raw = new_tbl["_gmv"]
    hot = hot.drop(columns=drop)
    new_tbl = new_tbl.drop(columns=[c for c in drop if c in new_tbl.columns])

    # 自社案件リスト (own_list.csv): 一致商品のリンク差し替え + 🎁付与
    from weekly_picks import norm_name
    own_rows = load_own_list()
    own_by_id = {r["商品ID"]: r for r in own_rows}
    own_by_name = {norm_name(r["商品名"]): r for r in own_rows}
    hot, n_own_hot = apply_own_list(hot, own_by_id, own_by_name)
    new_tbl, n_own_new = apply_own_list(new_tbl, own_by_id, own_by_name)
    if own_rows:
        print(f"[info] 自社案件リスト {len(own_rows)}件 / "
              f"リンク差替 売れ筋{n_own_hot}件・新商品{n_own_new}件")

    # ジャンルサマリー (件数と売上合計)
    def cat_summary(tbl, gmv_raw):
        s = tbl.copy()
        s["_g"] = gmv_raw.loc[s.index]
        agg = (s.groupby("ジャンル", sort=False)
                .agg(商品数=("商品名", "count"), 合計30日売上=("_g", "sum")))
        agg["合計30日売上"] = agg["合計30日売上"].map(fmt_money)
        return agg.reset_index()

    out_path = args.out or f"weekly_data_{dt.date.today():%Y%m%d}.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        cat_summary(hot, hot_gmv_raw).to_excel(w, sheet_name="カテゴリ別サマリー", index=False)
        hot.to_excel(w, sheet_name="売れ筋候補", index=False)
        new_tbl.to_excel(w, sheet_name="新商品候補", index=False)
        # 列幅調整
        for ws in w.book.worksheets:
            for col_cells in ws.columns:
                width = max(len(str(c.value)) for c in col_cells if c.value) + 2
                ws.column_dimensions[col_cells[0].column_letter].width = min(width, 50)

    print(f"[info] 売れ筋候補 {len(hot)}件 / 新商品候補 {len(new_tbl)}件 → {out_path}")

    if args.images:
        embed_images(out_path, ["売れ筋候補", "新商品候補"])

    if args.html or args.html_embed:
        from report_html import write_site, archive_site
        site_dir = "weekly_site"
        write_site(hot, new_tbl, site_dir, embed=args.html_embed, own_rows=own_rows)
        print(f"[info] Webサイト出力: {site_dir}/index.html (売れ筋), "
              f"{site_dir}/challenge.html (新商品)"
              + (" [画像埋め込み版]" if args.html_embed else ""))
        if not args.no_archive:
            arch_dir = archive_site(hot, new_tbl, root_dir=".", own_rows=own_rows)
            print(f"[info] アーカイブ保存: {arch_dir}/ と data/*.json "
                  f"(archive/ data/ もコミット対象)")

    # コンソール確認用の詳細ダンプ (件数・突合の目視確認向け。Discord投稿の内容とは別物で、
    # Discordとサイトには30日売上/単価/報酬率の3項目しか出さない)
    today = dt.date.today().strftime("%m/%d")
    lines = [f"# 📦 今週のおすすめ商品データ（{today}｜コンソール確認用）", "",
             "# 🔥 売れ筋候補"]
    for cat, g in hot.groupby("ジャンル", sort=False):
        lines += [f"## ▼ {cat}（{len(g)}件）"]
        for _, r in g.iterrows():
            lines += [
                f"**{r['商品名']}**（{r['カテゴリ']}）",
                f"　30日売上 {r['30日売上']}（{r['販売件数']}）｜直近7日動画売上 {r['直近7日動画売上']}（{r['直近7日動画本数']}本）",
                f"　単価 {r['単価']}｜報酬率 {r['報酬率']}→**1件{r['報酬目安/件']}**｜成立率 {r['成立率']}（参画{r['参画クリエイター数']}人）",
                f"　タグ: {r['タグ'] or 'なし'}", ""]
    lines += ["# 🚀 新商品候補"]
    for cat, g in new_tbl.groupby("ジャンル", sort=False):
        lines += [f"## ▼ {cat}（{len(g)}件）"]
        for _, r in g.iterrows():
            lines += [
                f"**{r['商品名']}**（{r['カテゴリ']}）",
                f"　成長率 {r['成長率']}｜30日売上 {r['30日売上']}｜掲載日 {r['掲載日']}",
                f"　単価 {r['単価']}｜報酬率 {r['報酬率']}｜成立率 {r['成立率']}",
                f"　タグ: {r['タグ'] or 'なし'}", ""]
    print("\n" + "\n".join(lines))

    if args.preview_post:
        post_discord_embeds(hot, new_tbl, "", args.top_per_cat,
                            preview_path=args.preview_post)
        print(f"[info] Discord投稿プレビューを書き出しました (送信はしていません): {args.preview_post}")

    if args.post:
        if not args.webhook:
            sys.exit("[error] --webhook または DISCORD_WEBHOOK_URL を設定してください")
        post_discord_embeds(hot, new_tbl, args.webhook, args.top_per_cat)
        print("[info] Discordに投稿しました")


if __name__ == "__main__":
    main()
